"""Excel workbook exporter for CareerSignal outputs."""

from __future__ import annotations

from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.config.schemas import Candidate, SkillTaxonomyConfig
from src.processing.skill_extractor import build_skill_patterns

ALL_JOBS_COLUMNS: list[tuple[str, str]] = [
    ("category_name", "Category"),
    ("match_tier", "Match Tier"),
    ("match_score", "Match Score"),
    ("job_title", "Job Title"),
    ("company", "Company"),
    ("industry", "Industry"),
    ("location", "Location"),
    ("location_normalized", "Normalized Location"),
    ("location_group", "Location Group"),
    ("work_arrangement", "Work Arrangement"),
    ("seniority", "Seniority"),
    ("employment_type", "Employment Type"),
    ("salary_range_text", "Salary Range"),
    ("salary_min", "Salary Min"),
    ("salary_max", "Salary Max"),
    ("salary_midpoint", "Salary Midpoint"),
    ("visa_signal", "Visa Signal"),
    ("visa_status", "Visa Status"),
    ("visa_evidence", "Visa Evidence"),
    ("visa_confidence", "Visa Confidence"),
    ("required_skills", "Required Skills"),
    ("preferred_skills", "Preferred Skills"),
    ("all_extracted_skills", "All Extracted Skills"),
    ("jd_post_link", "JD Post Link"),
    ("apply_link", "Apply Link"),
    ("date_posted", "Date Posted"),
    ("date_collected", "Date Collected"),
    ("source", "Source"),
    ("application_status", "Application Status"),
    ("reasoning_summary", "Reasoning Summary"),
]

LONG_TEXT_HEADERS = {
    "Required Skills",
    "Preferred Skills",
    "All Extracted Skills",
    "JD Post Link",
    "Apply Link",
    "Reasoning Summary",
    "Example Matching Job Titles",
    "Visa Signal Summary",
    "Visa Evidence",
}
CURRENCY_HEADERS = {
    "Salary Min",
    "Salary Max",
    "Salary Midpoint",
    "Average Salary Midpoint",
}
SCORE_HEADERS = {"Match Score", "Average Match Score", "Highest Match Score"}
LINK_HEADERS = {"JD Post Link", "Apply Link"}

MART_COLUMN_RENAMES: dict[str, str] = {
    "category_name": "Category",
    "jobs_found": "Jobs Found",
    "excellent_matches": "Excellent Matches",
    "strong_matches": "Strong Matches",
    "good_matches": "Good Matches",
    "average_match_score": "Average Match Score",
    "average_salary_midpoint": "Average Salary Midpoint",
    "remote_count": "Remote Count",
    "hybrid_count": "Hybrid Count",
    "onsite_count": "On-site Count",
    "unknown_work_arrangement_count": "Unknown Work Arrangement Count",
    "positive_visa_signal_count": "Positive Visa Signal Count",
    "negative_visa_signal_count": "Negative Visa Signal Count",
    "unknown_visa_signal_count": "Unknown Visa Signal Count",
    "match_tier": "Match Tier",
    "match_score": "Match Score",
    "job_title": "Job Title",
    "normalized_title": "Normalized Title",
    "company": "Company",
    "industry": "Industry",
    "location": "Location",
    "location_normalized": "Normalized Location",
    "location_group": "Location Group",
    "work_arrangement": "Work Arrangement",
    "seniority": "Seniority",
    "employment_type": "Employment Type",
    "salary_range_text": "Salary Range",
    "salary_min": "Salary Min",
    "salary_max": "Salary Max",
    "salary_midpoint": "Salary Midpoint",
    "visa_signal": "Visa Signal",
    "visa_status": "Visa Status",
    "visa_evidence": "Visa Evidence",
    "visa_confidence": "Visa Confidence",
    "required_skills": "Required Skills",
    "preferred_skills": "Preferred Skills",
    "all_extracted_skills": "All Extracted Skills",
    "jd_post_link": "JD Post Link",
    "apply_link": "Apply Link",
    "date_posted": "Date Posted",
    "date_collected": "Date Collected",
    "source": "Source",
    "application_status": "Application Status",
    "reasoning_summary": "Reasoning Summary",
    "skill": "Skill",
    "skill_group": "Skill Group",
    "appears_in_job_count": "Appears In Job Count",
    "appears_in_job_pct": "Appears In Job %",
    "in_candidate_profile": "In Candidate Profile",
    "gap_priority": "Gap Priority",
    "example_matching_job_titles": "Example Matching Job Titles",
    "matching_roles_count": "Matching Roles Count",
    "highest_match_score": "Highest Match Score",
    "best_matching_role": "Best Matching Role",
    "visa_signal_summary": "Visa Signal Summary",
    "priority": "Priority",
}


class ExcelExporter:
    """Builds a multi-tab Excel workbook for processed CareerSignal jobs."""

    def __init__(self, candidate: Candidate, taxonomy: SkillTaxonomyConfig) -> None:
        self.candidate = candidate
        self.taxonomy = taxonomy

    def export(
        self,
        jobs: list[dict[str, Any]],
        output_path: str | Path,
        top_match_threshold: float,
    ) -> Path:
        """Export processed jobs and summaries to an Excel workbook."""

        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)

        all_jobs_df = self._build_all_jobs_df(jobs)
        top_matches_df = self._build_top_matches_df(all_jobs_df, top_match_threshold)
        category_summary_df = self._build_category_summary_df(jobs)
        skill_gap_df = self._build_skill_gap_df(jobs)
        company_priority_df = self._build_company_priority_df(jobs)

        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            all_jobs_df.to_excel(writer, sheet_name="All Jobs", index=False)
            top_matches_df.to_excel(writer, sheet_name="Top Matches", index=False)
            category_summary_df.to_excel(writer, sheet_name="By Category Summary", index=False)
            skill_gap_df.to_excel(writer, sheet_name="Skill Gap Analysis", index=False)
            company_priority_df.to_excel(writer, sheet_name="Company Priority List", index=False)

            workbook = writer.book
            for worksheet in workbook.worksheets:
                self._format_worksheet(worksheet)

        return path

    def export_dataframes(
        self,
        sheets: dict[str, pd.DataFrame],
        output_path: str | Path,
    ) -> Path:
        """Export prebuilt tab dataframes, typically from dbt mart tables."""

        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)

        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            for sheet_name, dataframe in sheets.items():
                display_df = self._mart_display_df(dataframe)
                display_df.to_excel(writer, sheet_name=sheet_name, index=False)

            workbook = writer.book
            for worksheet in workbook.worksheets:
                self._format_worksheet(worksheet)

        return path

    def _mart_display_df(self, dataframe: pd.DataFrame) -> pd.DataFrame:
        if dataframe.empty:
            return dataframe.rename(columns=MART_COLUMN_RENAMES)
        display = dataframe.copy()
        for column in ("required_skills", "preferred_skills", "all_extracted_skills"):
            if column in display.columns:
                display[column] = display[column].map(self._json_list_to_text)
        return display.rename(columns=MART_COLUMN_RENAMES)

    def _json_list_to_text(self, value: Any) -> Any:
        if not isinstance(value, str):
            return value
        text = value.strip()
        if not text.startswith("["):
            return value
        try:
            import json

            parsed = json.loads(text)
        except json.JSONDecodeError:
            return value
        if isinstance(parsed, list):
            return ", ".join(str(item) for item in parsed)
        return value

    def _build_all_jobs_df(self, jobs: list[dict[str, Any]]) -> pd.DataFrame:
        rows: list[dict[str, Any]] = []
        for job in jobs:
            row: dict[str, Any] = {}
            for source_key, label in ALL_JOBS_COLUMNS:
                value = job.get(source_key)
                if isinstance(value, list):
                    value = ", ".join(str(item) for item in value)
                row[label] = value
            rows.append(row)
        return pd.DataFrame(rows, columns=[label for _, label in ALL_JOBS_COLUMNS])

    def _build_top_matches_df(
        self,
        all_jobs_df: pd.DataFrame,
        top_match_threshold: float,
    ) -> pd.DataFrame:
        if all_jobs_df.empty:
            return all_jobs_df.copy()

        top_matches = all_jobs_df[all_jobs_df["Match Score"] >= top_match_threshold].copy()
        if top_matches.empty:
            return top_matches

        top_matches["_date_posted_sort"] = pd.to_datetime(
            top_matches["Date Posted"], errors="coerce"
        )
        top_matches = top_matches.sort_values(
            by=["Match Score", "Salary Midpoint", "_date_posted_sort"],
            ascending=[False, False, False],
            na_position="last",
        )
        return top_matches.drop(columns=["_date_posted_sort"])

    def _build_category_summary_df(self, jobs: list[dict[str, Any]]) -> pd.DataFrame:
        columns = [
            "Category",
            "Jobs Found",
            "Excellent Matches",
            "Strong Matches",
            "Good Matches",
            "Average Match Score",
            "Average Salary Midpoint",
            "Remote Count",
            "Hybrid Count",
            "On-site Count",
            "Unknown Work Arrangement Count",
            "Positive Visa Signal Count",
            "Negative Visa Signal Count",
            "Unknown Visa Signal Count",
        ]
        if not jobs:
            return pd.DataFrame(columns=columns)

        rows: list[dict[str, Any]] = []
        jobs_by_category: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for job in jobs:
            jobs_by_category[str(job.get("category_name") or "Unknown")].append(job)

        for category, category_jobs in sorted(jobs_by_category.items()):
            tiers = Counter(job.get("match_tier") for job in category_jobs)
            work = Counter(job.get("work_arrangement") for job in category_jobs)
            visa = Counter(job.get("visa_signal") for job in category_jobs)
            salary_values = [
                float(job["salary_midpoint"])
                for job in category_jobs
                if job.get("salary_midpoint") is not None
            ]
            rows.append(
                {
                    "Category": category,
                    "Jobs Found": len(category_jobs),
                    "Excellent Matches": tiers["Excellent Match"],
                    "Strong Matches": tiers["Strong Match"],
                    "Good Matches": tiers["Good Match"],
                    "Average Match Score": round(
                        sum(float(job.get("match_score") or 0) for job in category_jobs)
                        / len(category_jobs),
                        1,
                    ),
                    "Average Salary Midpoint": round(
                        sum(salary_values) / len(salary_values), 2
                    )
                    if salary_values
                    else None,
                    "Remote Count": work["Remote"],
                    "Hybrid Count": work["Hybrid"],
                    "On-site Count": work["On-site"],
                    "Unknown Work Arrangement Count": work["Unknown"],
                    "Positive Visa Signal Count": visa["Positive"],
                    "Negative Visa Signal Count": visa["Negative"],
                    "Unknown Visa Signal Count": visa["Unknown"],
                }
            )
        return pd.DataFrame(rows, columns=columns)

    def _build_skill_gap_df(self, jobs: list[dict[str, Any]]) -> pd.DataFrame:
        columns = [
            "Skill",
            "Skill Group",
            "Appears In Job Count",
            "Appears In Job %",
            "In Candidate Profile",
            "Gap Priority",
            "Example Matching Job Titles",
        ]
        if not jobs:
            return pd.DataFrame(columns=columns)

        patterns = build_skill_patterns(self.candidate, self.taxonomy)
        pattern_lookup = {pattern.canonical.casefold(): pattern for pattern in patterns}
        skill_counts: Counter[str] = Counter()
        example_titles: dict[str, list[str]] = defaultdict(list)

        for job in jobs:
            job_skills = {
                str(skill)
                for skill in job.get("all_extracted_skills", [])
                if str(skill).strip()
            }
            for skill in job_skills:
                skill_counts[skill] += 1
                if len(example_titles[skill]) < 3:
                    example_titles[skill].append(str(job.get("job_title") or "Unknown"))

        rows: list[dict[str, Any]] = []
        total_jobs = max(len(jobs), 1)
        for skill, count in skill_counts.most_common():
            pattern = pattern_lookup.get(skill.casefold())
            in_candidate = bool(pattern.in_candidate_profile) if pattern else False
            skill_group = pattern.skill_group if pattern else "Market Skill"
            appears_percent = count / total_jobs

            if not in_candidate and (appears_percent >= 0.35 or count >= 4):
                priority = "High"
            elif not in_candidate and count >= 2:
                priority = "Medium"
            else:
                priority = "Low"

            rows.append(
                {
                    "Skill": skill,
                    "Skill Group": skill_group,
                    "Appears In Job Count": count,
                    "Appears In Job %": round(appears_percent, 3),
                    "In Candidate Profile": "Yes" if in_candidate else "No",
                    "Gap Priority": priority,
                    "Example Matching Job Titles": ", ".join(example_titles[skill]),
                }
            )
        return pd.DataFrame(rows, columns=columns)

    def _build_company_priority_df(self, jobs: list[dict[str, Any]]) -> pd.DataFrame:
        columns = [
            "Company",
            "Industry",
            "Matching Roles Count",
            "Average Match Score",
            "Highest Match Score",
            "Average Salary Midpoint",
            "Best Matching Role",
            "Visa Signal Summary",
            "Priority",
        ]
        if not jobs:
            return pd.DataFrame(columns=columns)

        rows: list[dict[str, Any]] = []
        jobs_by_company: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for job in jobs:
            jobs_by_company[str(job.get("company") or "Unknown")].append(job)

        for company, company_jobs in sorted(jobs_by_company.items()):
            salary_values = [
                float(job["salary_midpoint"])
                for job in company_jobs
                if job.get("salary_midpoint") is not None
            ]
            score_values = [float(job.get("match_score") or 0) for job in company_jobs]
            best_job = max(company_jobs, key=lambda job: float(job.get("match_score") or 0))
            tiers = Counter(job.get("match_tier") for job in company_jobs)
            visa = Counter(job.get("visa_signal") for job in company_jobs)

            if tiers["Excellent Match"] >= 1 or tiers["Strong Match"] >= 2:
                priority = "High"
            elif tiers["Strong Match"] >= 1 or tiers["Good Match"] >= 1:
                priority = "Medium"
            else:
                priority = "Low"

            visa_summary = "; ".join(
                f"{signal}: {count}" for signal, count in sorted(visa.items()) if signal
            )

            rows.append(
                {
                    "Company": company,
                    "Industry": best_job.get("industry") or "Unknown",
                    "Matching Roles Count": len(company_jobs),
                    "Average Match Score": round(sum(score_values) / len(score_values), 1),
                    "Highest Match Score": round(max(score_values), 1),
                    "Average Salary Midpoint": round(
                        sum(salary_values) / len(salary_values), 2
                    )
                    if salary_values
                    else None,
                    "Best Matching Role": best_job.get("job_title"),
                    "Visa Signal Summary": visa_summary,
                    "Priority": priority,
                }
            )
        return pd.DataFrame(rows, columns=columns)

    def _format_worksheet(self, worksheet: Worksheet) -> None:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(bold=True, color="FFFFFF")

        headers = {
            cell.column: str(cell.value)
            for cell in worksheet[1]
            if cell.value is not None
        }
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for column_idx, header in headers.items():
            column_letter = get_column_letter(column_idx)
            width = self._recommended_width(worksheet, column_idx, header)
            worksheet.column_dimensions[column_letter].width = width

            if header in CURRENCY_HEADERS:
                for cell in worksheet[column_letter][1:]:
                    cell.number_format = '$#,##0'
            elif header in SCORE_HEADERS:
                for cell in worksheet[column_letter][1:]:
                    cell.number_format = "0.0"
                self._add_score_conditional_formatting(worksheet, column_letter)
            elif header == "Appears In Job %":
                for cell in worksheet[column_letter][1:]:
                    cell.number_format = "0.0%"

            if header in LONG_TEXT_HEADERS:
                for cell in worksheet[column_letter][1:]:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

            if header in LINK_HEADERS:
                self._format_hyperlinks(worksheet, column_letter)

    def _recommended_width(self, worksheet: Worksheet, column_idx: int, header: str) -> int:
        max_length = len(header)
        for row_idx in range(2, min(worksheet.max_row, 200) + 1):
            value = worksheet.cell(row=row_idx, column=column_idx).value
            if value is not None:
                max_length = max(max_length, len(str(value)))

        if header in LONG_TEXT_HEADERS:
            return min(max(max_length + 2, 24), 70)
        return min(max(max_length + 2, 12), 34)

    def _format_hyperlinks(self, worksheet: Worksheet, column_letter: str) -> None:
        for cell in worksheet[column_letter][1:]:
            if isinstance(cell.value, str) and cell.value.startswith(("http://", "https://")):
                cell.hyperlink = cell.value
                cell.style = "Hyperlink"

    def _add_score_conditional_formatting(self, worksheet: Worksheet, column_letter: str) -> None:
        if worksheet.max_row < 2:
            return

        cell_range = f"{column_letter}2:{column_letter}{worksheet.max_row}"
        excellent_fill = PatternFill("solid", fgColor="63BE7B")
        strong_fill = PatternFill("solid", fgColor="9DD7A7")
        good_fill = PatternFill("solid", fgColor="FFEB84")
        low_fill = PatternFill("solid", fgColor="F4B183")

        worksheet.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="greaterThanOrEqual", formula=["90"], fill=excellent_fill),
        )
        worksheet.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="between", formula=["80", "89.999"], fill=strong_fill),
        )
        worksheet.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="between", formula=["70", "79.999"], fill=good_fill),
        )
        worksheet.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="lessThan", formula=["70"], fill=low_fill),
        )
