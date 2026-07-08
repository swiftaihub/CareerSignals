from __future__ import annotations

from openpyxl import load_workbook

from src.config.loader import load_configs
from src.exporters.excel_exporter import ExcelExporter


def test_excel_exporter_creates_required_sheets(tmp_path) -> None:
    configs = load_configs(".")
    output_path = tmp_path / "job_search_tracker.xlsx"
    jobs = [
        {
            "category_name": "Data Scientist - Tech",
            "match_tier": "Excellent Match",
            "match_score": 94.5,
            "job_title": "Product Data Scientist",
            "company": "NimbusAI",
            "industry": "Technology",
            "location": "Remote",
            "work_arrangement": "Remote",
            "seniority": "Senior",
            "employment_type": "full-time",
            "salary_range_text": "$130k-$170k",
            "salary_min": 130000,
            "salary_max": 170000,
            "salary_midpoint": 150000,
            "visa_signal": "Positive",
            "required_skills": ["Python", "SQL"],
            "preferred_skills": ["LLM"],
            "all_extracted_skills": ["Python", "SQL", "LLM", "Airflow"],
            "jd_post_link": "https://jobs.example.com/nimbusai/product-data-scientist-ai-platform",
            "apply_link": "https://jobs.example.com/nimbusai/product-data-scientist-ai-platform/apply",
            "date_posted": "2026-07-01",
            "date_collected": "2026-07-07",
            "source": "MockBoard",
            "application_status": "Not Applied",
            "reasoning_summary": "Excellent match with strong skills alignment.",
        }
    ]

    ExcelExporter(
        configs.candidate_profile.candidate,
        configs.skill_taxonomy,
    ).export(jobs, output_path, top_match_threshold=80)

    workbook = load_workbook(output_path)

    assert workbook.sheetnames == [
        "All Jobs",
        "Top Matches",
        "By Category Summary",
        "Skill Gap Analysis",
        "Company Priority List",
    ]
    assert workbook["All Jobs"].max_row == 2
